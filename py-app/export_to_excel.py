"""
Excel Export Utility for Restaurant Recommender
Creates Excel files with restaurant data, reviews, and analytics
"""

import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

import pandas as pd
import sqlite3
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def export_restaurant_data_to_excel():
    """
    Export all restaurant data to a comprehensive Excel file
    """
    print("🍽️ Creating Excel Export of Restaurant Data...")
    
    # Get database path
    db_path = os.path.join(os.path.dirname(__file__), '..', 'austin_restaurants.db')
    
    # Connect to database
    conn = sqlite3.connect(db_path)
    
    try:
        # Load all data
        print("📊 Loading data from database...")
        
        # Restaurants with details
        restaurants_query = """
        SELECT 
            r.id,
            r.name,
            r.address,
            r.rating,
            r.price_level,
            r.user_ratings_total,
            r.lat,
            r.lng,
            r.place_tags as cuisine,
            COALESCE(pd.serves_vegetarian_food, 0) as serves_vegetarian,
            COALESCE(pd.dine_in, 0) as dine_in,
            COALESCE(pd.takeout, 0) as takeout,
            COALESCE(pd.delivery, 0) as delivery,
            COALESCE(pd.serves_beer, 0) as serves_beer,
            COALESCE(pd.serves_wine, 0) as serves_wine,
            COALESCE(pd.serves_breakfast, 0) as serves_breakfast,
            COALESCE(pd.serves_brunch, 0) as serves_brunch,
            COALESCE(pd.serves_lunch, 0) as serves_lunch,
            COALESCE(pd.serves_dinner, 0) as serves_dinner,
            pd.website,
            pd.formatted_phone_number,
            pd.business_status
        FROM restaurants r
        LEFT JOIN place_details pd ON r.id = pd.id
        WHERE r.business_status = 'OPERATIONAL'
        """
        restaurants = pd.read_sql_query(restaurants_query, conn)
        
        # Reviews
        reviews_query = """
        SELECT 
            id as restaurant_id,
            author_name as author,
            rating,
            text as review_text,
            time as review_date
        FROM reviews
        WHERE text IS NOT NULL AND text != ''
        """
        reviews = pd.read_sql_query(reviews_query, conn)
        
        # Convert timestamps
        reviews['review_date'] = pd.to_datetime(reviews['review_date'], unit='s')
        
        # Convert boolean columns
        bool_cols = ['serves_vegetarian', 'dine_in', 'takeout', 'delivery', 
                     'serves_beer', 'serves_wine', 'serves_breakfast', 
                     'serves_brunch', 'serves_lunch', 'serves_dinner']
        for col in bool_cols:
            if col in restaurants.columns:
                restaurants[col] = restaurants[col].astype(bool)
        
        # Create Excel file
        excel_filename = f"austin_restaurants_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_path = os.path.join(os.path.dirname(__file__), excel_filename)
        
        print(f"📝 Creating Excel file: {excel_filename}")
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet 1: Restaurant Overview
            restaurants.to_excel(writer, sheet_name='Restaurants', index=False)
            
            # Sheet 2: Reviews
            reviews.to_excel(writer, sheet_name='Reviews', index=False)
            
            # Sheet 3: Analytics Summary
            analytics_data = {
                'Metric': [
                    'Total Restaurants',
                    'Total Reviews',
                    'Average Rating',
                    'Restaurants with Price Info',
                    'Vegetarian Options',
                    'Delivery Available',
                    'Takeout Available',
                    'Serves Beer',
                    'Serves Wine'
                ],
                'Count': [
                    len(restaurants),
                    len(reviews),
                    f"{restaurants['rating'].mean():.2f}",
                    restaurants['price_level'].notna().sum(),
                    restaurants['serves_vegetarian'].sum(),
                    restaurants['delivery'].sum(),
                    restaurants['takeout'].sum(),
                    restaurants['serves_beer'].sum(),
                    restaurants['serves_wine'].sum()
                ]
            }
            analytics_df = pd.DataFrame(analytics_data)
            analytics_df.to_excel(writer, sheet_name='Analytics', index=False)
            
            # Sheet 4: Top Rated Restaurants
            top_rated = restaurants.nlargest(20, 'rating')[['name', 'address', 'rating', 'user_ratings_total', 'cuisine']]
            top_rated.to_excel(writer, sheet_name='Top Rated', index=False)
        
        # Format the Excel file
        print("🎨 Formatting Excel file...")
        format_excel_file(excel_path)
        
        print(f"✅ Excel file created successfully: {excel_path}")
        print(f"📊 Data exported:")
        print(f"   - {len(restaurants)} restaurants")
        print(f"   - {len(reviews)} reviews")
        print(f"   - 4 sheets: Restaurants, Reviews, Analytics, Top Rated")
        
        return excel_path
        
    finally:
        conn.close()

def format_excel_file(excel_path):
    """
    Format the Excel file with headers, colors, and styling
    """
    workbook = openpyxl.load_workbook(excel_path)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Format each sheet
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Format headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(excel_path)

def export_search_results_to_excel(query, results_df):
    """
    Export search results to Excel
    """
    print(f"🔍 Exporting search results for '{query}' to Excel...")
    
    # Create filename with query
    safe_query = "".join(c for c in query if c.isalnum() or c in (' ', '-', '_')).rstrip()
    excel_filename = f"search_results_{safe_query}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    excel_path = os.path.join(os.path.dirname(__file__), excel_filename)
    
    # Export to Excel with formatting
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='Search Results', index=False)
    
    format_excel_file(excel_path)
    
    print(f"✅ Search results exported: {excel_path}")
    return excel_path

if __name__ == "__main__":
    try:
        excel_file = export_restaurant_data_to_excel()
        print(f"\n🎉 Success! Excel file created: {excel_file}")
        print(f"📁 Location: {os.path.dirname(excel_file)}")
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        import traceback
        traceback.print_exc()


